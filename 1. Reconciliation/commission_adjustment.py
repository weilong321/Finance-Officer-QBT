import os
from glob import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import utils

def append_all(dfs, names, df, name):
    dfs.append(df)
    names.append(name)
    return

def add_sheets(dfs, names):
    path = "C:\\Users\\ctm_mchen\\OneDrive - Helloworld Travel Ltd\\Desktop\\reconciliation\\Commission Adjustments.xlsx"
    wb = Workbook()
    ws = None
    is_first = True
    for i in range(len(names)):
        if is_first:
            is_first = False
            ws = wb.worksheets[0]
            ws.title = names[i]
        else:
            ws = wb.create_sheet(names[i])
        for col_idx, value in enumerate(dfs[i].columns.tolist()):
            ws.cell(row=1, column=col_idx+1).value = value
        for row_idx, row in dfs[i].iterrows():
            for col_idx, cell_value in enumerate(row.tolist()):
                try:
                    rounded_value = round(float(cell_value), 2)
                    ws.cell(row=row_idx+2, column=col_idx+1).value = rounded_value
                except:
                    ws.cell(row=row_idx+2, column=col_idx+1).value = cell_value
        ws.freeze_panes = 'A2'
        for column in ws.columns:
            for cell in column:
                column_letter = utils.get_column_letter(cell.column)
                ws.column_dimensions[column_letter].auto_size = True
    wb.save(path)


def get_files_dates():
    folder_path = 'C:/Users/ctm_mchen/OneDrive - Helloworld Travel Ltd/Desktop/reconciliation/2021 Reconciliation'
    files = []
    dates = []
    subdirs = [x[0] for x in os.walk(folder_path)]
    for subdir in subdirs:
        files.extend(glob(subdir + '/Coverpage*'))
        dates.extend(glob(subdir + '/* Opening.xlsx'))
    return files, dates

def summary_sheet(file, dfs, names):
    week_ending = []
    small_variance = []
    airticket_fee = []
    commissions = []
    ongoing_diff = []

    for file in files:
        df = pd.read_excel(open(file, 'rb'), sheet_name='Coverpage')
        week_ending.append(df.iat[0, 3].date())
        small_variance.append(df.iat[30, 2])
        airticket_fee.append(df.iat[32, 2])
        commissions.append(df.iat[34, 2])
        ongoing_diff.append(df.iat[58, 2])
    
    data = list(zip(week_ending, small_variance, airticket_fee, commissions, ongoing_diff))
    df = pd.DataFrame(data, columns=['Week Ending', 'Small Variance', 'Airticket Fee', 'Commissions', 'Ongoing Difference'])
    df = df.sort_values(by='Week Ending', ascending=False)
    append_all(dfs, names, df, "Summary")

    return df

def commission_sheet(files, dates, dfs, names):
    df = pd.DataFrame()
    is_first = True
    for file, date in zip(files, dates):
        formatted_date = date.split("\\")[-1].split(" ")[0]
        if is_first:
            df = pd.read_excel(open(file, 'rb'), sheet_name='Commission Only')
            df = df.assign(Week=formatted_date)
            is_first = False
            continue
        commission = pd.read_excel(open(file, 'rb'), sheet_name='Commission Only').assign(Week=formatted_date)
        df = pd.concat([df, commission])
    df = df.reset_index(drop=True)

    groups = df.groupby(by=abs(df['Nett Paid (DIFF)']), as_index=False)
    for name, group in groups:
        if group["Nett Paid (DIFF)"].sum().round(2) == 0:
            df = df.drop(group.index)
    df = df.sort_values(by=['Nett Paid (DIFF)'], key=abs, ascending=False).reset_index(drop=True)

    append_all(dfs, names, df, "Sus Commissions")
    return df



# MAIN #
dfs = []
names = []

files, dates = get_files_dates()
summary = summary_sheet(files, dfs, names)
commission = commission_sheet(files, dates, dfs, names)
add_sheets(dfs, names)