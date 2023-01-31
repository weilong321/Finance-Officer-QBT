import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import utils
import os

def get_week_ending():
    while True:
        week_ending = input("""\tEnter the week ending date in this format:\n
        \t\"YYYY.MM.DD\" (e.g. 2021.05.02)\n""")
        confirm = input("\tYou have selected {}.\n\tIs this the date you want?\n\tType y for yes and anything else for no.\n".format(week_ending))
        if confirm != "y":
            continue
        return week_ending

def get_bsp_report_date():
    while True:
        bsp_report_date = input("""\tEnter the BSP Report date in this format: \n
        \t\t\"YYYYMMDD\" (e.g. 20220930)\n""")
        confirm = input("\tYou have selected {}.\n\tIs this the date you want?\n\tType y for yes and anything else for no.\n".format(bsp_report_date))
        if confirm != "y":
            continue
        return bsp_report_date

def get_first_sheet(wb):
    global is_first
    is_first = False
    return wb.worksheets[0]

def add_sheets(path, dfs, names):
    wb = load_workbook(path)
    ws = None
    is_first = True
    for i in range(len(names)):
        if is_first:
            is_first = False
            ws = wb.worksheets[0]
            ws.title = names[i]
        else:
            ws = wb.create_sheet(names[i])
        for col_idx, value in enumerate(dfs[i].columns.tolist()):
            ws.cell(row=1, column=col_idx+1).value = value
        for row_idx, row in dfs[i].iterrows():
            for col_idx, cell_value in enumerate(row.tolist()):
                try:
                    rounded_value = round(float(cell_value), 2)
                    ws.cell(row=row_idx+2, column=col_idx+1).value = rounded_value
                except:
                    ws.cell(row=row_idx+2, column=col_idx+1).value = cell_value
        ws.freeze_panes = 'A2'
        for column in ws.columns:
            for cell in column:
                column_letter = utils.get_column_letter(cell.column)
                ws.column_dimensions[column_letter].auto_size = True
    wb.save(path)
    dfs.clear()
    names.clear()
    return

def create_new_workbook(path):
    wb = Workbook()
    wb.save(path)

def append_all(dfs, names, df, name):
    dfs.append(df)
    names.append(name)
    return

def get_week(dfs, names, bsp_report_date):
    while True:
        week_ending = input("\tEnter the week (sheet name):\n")
        confirm = input("\tYou have selected {}.\n\tIs this the week you want?\n\tType y for yes and anything else for no.\n".format(week_ending))
        if confirm != "y":
            continue
        try:
            df1 = pd.read_excel(open('Weekly Consolidated Stmnt 28Jun21 to 26Jun22.xlsx', 'rb'), sheet_name=week_ending)
            df2 = pd.read_excel(open('BSP Report - {}.xlsm'.format(bsp_report_date), 'rb'), sheet_name="Clean", header = 1)
            df3 = pd.read_excel(open('BSP-AGM Reconcilliation.xlsx', 'rb'), sheet_name="Reconcilliation")
            nettpaid = df3.iat[10, 2]
            sum1 = df2[['Sales Price', 'GST']].sum().sum()
            sum2 = df2[['Commission', 'UATP']].sum().sum()
            air_nett_payment = df1[[' Nett Due']].sum().iloc[0]
            bsp_nett_payment = sum1 - sum2
            check = bsp_nett_payment - nettpaid
            print("\t\t${} diff between actual BSP posting and raw data report used.".format(check.round(2)))
            confirm = input("\t\tIs this difference acceptable?\n\t\tType 'y' for yes, anything else for no.\n")
            if confirm != "y":
                return None, None, None, None, None, None
            path = 'Raw data for the week {}.xlsx'.format(week_ending)
            append_all(dfs, names, df1, "Airticket statement")
            append_all(dfs, names, df2, "BSP Source data")
            create_new_workbook(path)
            print("New file has been created under the name ({}.csv)!".format(week_ending))
            return week_ending, path, -bsp_nett_payment.round(2), air_nett_payment.round(2), df1, df2
        except:
            print("\t\'{}\' is not a valid sheet in the Airticket Statements.".format(week_ending))
            try_again = input("\tWould you like to try again?\n\tPress y for yes and anything else for no.\n")
            if try_again != "y":
                print("\tProgram will now stop.")
                return None, None, None, None, None, None

def drop_rows(main, df):
    return main.drop(df.index)

def reset_indices(dfs, df):
    df = df.reset_index(drop=True)
    dfs.append(df)
    return

def zip_tickets_sources(df, column_name, week_end_date, bsp_or_air):
    tickets = df.loc[:, column_name].to_numpy()
    data_source = np.array([], dtype='object')
    for i in range(len(tickets)):
        if bsp_or_air == 0:
            data_source = np.append(data_source, "BSP Source data {}".format(week_end_date))
        else:
            data_source = np.append(data_source, "Airticket statement {}".format(week_end_date))
    combined = np.column_stack((tickets, data_source))
    return combined

def get_cols(df, bsp_or_air):
    if bsp_or_air == 0:
        cols = df[['Ticket', 'Fare Price', 'Airport Tax', 'Sales Price', 'GST', 'Commission', 'UATP']]
        return cols
    elif bsp_or_air == 1:
        cols = df[['Ticket Number', ' Fare Credit', ' OB Fee', ' Tax', ' GST on OB Fee', ' Fees', ' GST amount', ' GST on Commission', ' Commission', ' Total Credit', ' Nett Due']]
        cols.columns = ['Ticket', 'Fare Credit', 'OB Fee', 'Tax', 'GST on OB Fee', 'Fees', 'GST amount', 'GST on Commission', 'Commission', 'Total Credit', 'Nett Due']
        return cols
    else:
        cols = df[['Ticket Number', 'Data Source']]
        cols.columns = ['Ticket', 'Data Source']
        return cols

def get_no_dupe_sheet(dfs, names, bspticket, airticket):
    bsp_combined = zip_tickets_sources(bspticket, "Ticket", week_end_date, 0)
    air_combined = zip_tickets_sources(airticket, "Ticket Number", week_end_date, 1)
    all_combined = np.concatenate((bsp_combined, air_combined), axis = 0)
    no_duplicates = pd.DataFrame(all_combined, columns=["Ticket Number", "Data Source"]).drop_duplicates(subset=["Ticket Number"], keep="first").reset_index(drop=True)
    append_all(dfs, names, no_duplicates, "Removed Duplicates")
    return no_duplicates, bspticket, airticket

def get_reconciliation_sheet(data_source, dfs, names, bsp, air):
    bsp_cleaned = get_cols(bsp, 0).groupby(["Ticket"]).sum()
    air_cleaned = get_cols(air, 1).groupby(["Ticket"]).sum()
    data_source_col = get_cols(data_source, 2)
    bsp_cleaned = bsp_cleaned.assign(Nett_Paid = lambda x: x['Sales Price'] + x['GST'] - x['Commission'] - x['UATP'])
    bsp_cleaned.columns = ['Fare Price (BSP)', 'Airport Tax (BSP)', 'Sales Price (BSP)', 'GST (BSP)', 'Commission (BSP)', 'UATP (BSP)', 'Nett Paid (BSP)']
    air_cleaned = air_cleaned.assign(Fare_Price = lambda x: x['Fare Credit'] + x['OB Fee'],
                                    GST = lambda x: x['GST amount'] + x['GST on OB Fee'] + x['GST on Commission'],
                                    Airport_Tax = lambda x: x['Tax'] + x['OB Fee'] + x['GST on OB Fee'] - x['GST'],
                                    Sales_Price = lambda x: x['Fare_Price'] + x['Airport_Tax'])
    air_cleaned = air_cleaned[['Fare_Price', 'Airport_Tax', 'Fees', 'Sales_Price', 'GST', 'Commission', 'Total Credit', 'Nett Due']]
    air_cleaned.columns = ['Fare Price (AIR)', 'Airport Tax (AIR)', 'Fees (AIR)', 'Sales Price (AIR)', 'GST (AIR)', 'Commission (AIR)', 'UATP (AIR)', 'Nett Paid (AIR)']
    final_cleaned = pd.merge(bsp_cleaned, air_cleaned, on='Ticket', how='outer').fillna(0)
    final_cleaned = final_cleaned.assign(fareprice = lambda x: x['Fare Price (BSP)'] - x['Fare Price (AIR)'],
                                        airporttax = lambda x: x['Airport Tax (BSP)'] - x['Airport Tax (AIR)'],
                                        salesprice = lambda x: x['Sales Price (BSP)'] - x['Sales Price (AIR)'],
                                        gst = lambda x: x['GST (BSP)'] - x['GST (AIR)'],
                                        commission = lambda x: x['Commission (BSP)'] - x['Commission (AIR)'],
                                        UATP = lambda x: x['UATP (BSP)'] - x['UATP (AIR)'],
                                        nettpaid = lambda x: x['Nett Paid (BSP)'] - x['Nett Paid (AIR)'],
                                        variance = lambda x: x['UATP (BSP)'] + x['Nett Paid (BSP)'] - x['UATP (AIR)'] - x['Nett Paid (AIR)'])
    final_cleaned.columns = ['Fare Price (BSP)', 'Airport Tax (BSP)', 'Sales Price (BSP)', 'GST (BSP)', 'Commission (BSP)', 'UATP (BSP)', 'Nett Paid (BSP)',
                            'Fare Price (AIR)', 'Airport Tax (AIR)', 'Fees (AIR)', 'Sales Price (AIR)', 'GST (AIR)', 'Commission (AIR)', 'UATP (AIR)', 'Nett Paid (AIR)',
                            'Fare Price (DIFF)', 'Airport Tax (DIFF)', 'Sales Price (DIFF)', 'GST (DIFF)', 'Commission (DIFF)', 'UATP (DIFF)', 'Nett Paid (DIFF)', 'Variance (DIFF)']
    final_cleaned = final_cleaned.round(2).reset_index().rename(columns={'index': 'Ticket'})
    final_cleaned = final_cleaned.merge(data_source_col, on='Ticket')
    column_order = final_cleaned.columns.tolist()
    last_column = column_order.pop()
    column_order.insert(1, last_column)
    final_cleaned = final_cleaned.reindex(columns=column_order)
    append_all(dfs, names, final_cleaned, "Reconciliation")
    return final_cleaned

def split_reconciliation(reconciliation, dfs, names):
    names_ls = ['UATP Tickets', 'Matches', 'Current Timing Difference', 'AIRTICKET FEES', 'Commission Only', 'Small Variances']
    dfs_ls = []
    # get uatp tickets
    uatp = reconciliation.loc[(reconciliation['Nett Paid (BSP)'] == 0) & (reconciliation['Nett Paid (AIR)'] == 0)]
    reconciliation = drop_rows(reconciliation, uatp)
    reset_indices(dfs_ls, uatp)
    # Get matches
    match = reconciliation.loc[(reconciliation['Nett Paid (DIFF)'] == 0)]
    reconciliation = drop_rows(reconciliation, match)
    reset_indices(dfs_ls, match)
    # get current timing difference
    bsp_diff = reconciliation[(reconciliation.iloc[:,2:9] == 0).all(axis=1)]
    air_diff = reconciliation[(reconciliation.iloc[:,9:17] == 0).all(axis=1)]
    curr_time_diff = pd.concat([bsp_diff, air_diff])
    reconciliation = drop_rows(reconciliation, curr_time_diff)
    reset_indices(dfs_ls, curr_time_diff)
    # Get airticket fees
    airticket_fees = reconciliation.loc[(reconciliation['Fees (AIR)'] != 0)]
    reconciliation = drop_rows(reconciliation, airticket_fees)
    reset_indices(dfs_ls, airticket_fees)
    total_air_fee = airticket_fees['Nett Paid (DIFF)'].sum().round(2)
    # Get Commission only
    commission = reconciliation.loc[(reconciliation['Commission (DIFF)'] + reconciliation['Nett Paid (DIFF)'] == 0)]
    reconciliation = drop_rows(reconciliation, commission)
    reset_indices(dfs_ls, commission)
    total_commission = commission['Nett Paid (DIFF)'].sum().round(2)
    # Get small variances
    small_variances = reconciliation
    reconciliation = drop_rows(reconciliation, small_variances)
    reset_indices(dfs_ls, small_variances)
    total_var = small_variances['Nett Paid (DIFF)'].sum().round(2)

    for elem in names_ls:
        names.append(elem)
    for elem in dfs_ls:
        dfs.append(elem)

    return curr_time_diff, small_variances, total_air_fee, total_commission, total_var

def combine_timing_diff(week_end_date, curr_time_diff, dfs, names):
    source = "{} Opening.xlsx".format(week_end_date)
    ongoing_time_diff = pd.read_excel(open(source, 'rb'), sheet_name="Ongoing Timing Difference")
    total_ongoing_diff = -ongoing_time_diff[['Nett Paid (DIFF)']].sum().round(2).values[0]
    print("Ongoing timing diff is: {}".format(total_ongoing_diff))
    if ongoing_time_diff.shape[1] == 30:
        ongoing_time_diff = ongoing_time_diff.iloc[:, :-5].reset_index(drop=True)
    else:
        ongoing_time_diff = ongoing_time_diff.reset_index(drop=True)
    ongoing_time_diff.columns = ['Ticket', 'Data Source',
                             'Fare Price (BSP)', 'Airport Tax (BSP)', 'Sales Price (BSP)', 'GST (BSP)', 'Commission (BSP)', 'UATP (BSP)', 'Nett Paid (BSP)',
                             'Fare Price (AIR)', 'Airport Tax (AIR)', 'Fees (AIR)', 'Sales Price (AIR)', 'GST (AIR)', 'Commission (AIR)', 'UATP (AIR)', 'Nett Paid (AIR)',
                             'Fare Price (DIFF)', 'Airport Tax (DIFF)', 'Sales Price (DIFF)', 'GST (DIFF)', 'Commission (DIFF)', 'UATP (DIFF)', 'Nett Paid (DIFF)', 'Variance (DIFF)']
    ongoing_time_diff = pd.concat([ongoing_time_diff, curr_time_diff], axis=0, ignore_index=True)
    return ongoing_time_diff

def process_timing_diff(ongoing_time_diff, small_var, dfs, names):
    no_dupes = ongoing_time_diff.drop_duplicates(subset="Ticket", keep=False)
    dropped = pd.DataFrame()
    temp = ongoing_time_diff[~ongoing_time_diff.index.isin(no_dupes.index)]
    dropped = pd.concat([dropped, temp])
    # time differences matched
    temp_df = pd.DataFrame()
    grouped = dropped.groupby('Ticket')
    for name, group in grouped:
        if group["Nett Paid (DIFF)"].sum().round(2) == 0:
            temp_df = pd.concat([temp_df, group])
    time_diff_match = temp_df.reset_index(drop=True)
    append_all(dfs, names, time_diff_match, "Timing Difference Matched")
    # Move small differences
    small_diff = drop_rows(dropped, temp_df).reset_index(drop=True)
    combined_small = pd.concat([small_var, small_diff]).reset_index(drop=True)
    combined_small = combined_small.sort_values(by=['Nett Paid (DIFF)'], key=abs, ascending=False).reset_index(drop=True)
    append_all(dfs, names, combined_small, "Small Variance")
    # Matching all airtickets to the most previous BSP week
    date_sources = no_dupes['Data Source'].unique()
    earliest_source = np.array([date_sources[0]])
    for source in date_sources:
        temp = source.split(" ")
        if temp[0] == "BSP" and earliest_source[0].split(" ")[-1] > temp[-1]:
            earliest_source[0] = source
        if temp[0] == "Airticket" and source not in earliest_source:
            earliest_source = np.append(earliest_source, source)
    filtered = no_dupes.loc[no_dupes['Data Source'].isin(earliest_source)]
    carryover_diff = drop_rows(no_dupes, filtered).reset_index(drop=True)
    filtered = filtered.sort_values(by=['Nett Paid (DIFF)'], key=abs, ascending=False).reset_index(drop=True)
    append_all(dfs, names, filtered, "Remaining time diff")
    append_all(dfs, names, carryover_diff, "Carryover time diff")
    total_ongoing_diff = carryover_diff['Nett Paid (DIFF)'].sum().round(2)
    return total_ongoing_diff

def coversheet_name_change(week):
    new_path = 'Coverpage for {}.xlsx'.format(week)
    os.rename('COVERPAGE TEMPLATE.xlsx', new_path)
    return
    
########################################
            #   MAIN   #
########################################
dfs = []
names = []

week_end_date = get_week_ending()
bsp_report_date = get_bsp_report_date()

week, path, bsp_nett_payment, air_nett_payment, airticket, bspticket = get_week(dfs, names, bsp_report_date)
if week == None:
    exit()

# No duplicates + get source
no_duplicates, bsp, air = get_no_dupe_sheet(dfs, names, bspticket, airticket)

# Get reconciliation sheet
reconciliation = get_reconciliation_sheet(no_duplicates, dfs, names, bsp, air)

# Split reconciliation sheet
curr_time_diff, small_variances, total_air, total_commission, total_var = split_reconciliation(reconciliation, dfs, names)

# combine ongoing and current timing difference
ongoing_time_diff = combine_timing_diff(week_end_date, curr_time_diff, dfs, names)

# Process timing difference
total_ongoing_diff = process_timing_diff(ongoing_time_diff, small_variances, dfs, names)

i = names.index("Small Variances")
dfs.pop(i)
names.pop(i)

print("Total bsp nett due is: {}".format(bsp_nett_payment))
print("Total airticket nett due is: {}".format(air_nett_payment))

# Add all sheets to excel
add_sheets(path, dfs, names)

coversheet_name_change(week)