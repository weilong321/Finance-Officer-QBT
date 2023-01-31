# Load first sheet of excel document 
import openpyxl

# Import required libraries + chromedriver
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
import os
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains

# Stop window from closing
from selenium.webdriver.chrome.options import Options
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)

TABLE_XPATH = '/html/body/agm-root/div/agm-dynamic/agm-window[3]/div/div[2]/div/div/div/form/agm-dynamic/agm-panel[1]/div/fieldset/agm-dynamic/agm-window-container/div/div[1]/agm-tab-content/fieldset/agm-dynamic/agm-window[2]/div/div[2]/div/div/div[2]/form/agm-dynamic/agm-grid/div/div[1]/table/tbody/'
XPATH_VAR = ['tr[1]/td[3]', 'tr[2]/td[3]', 'tr[1]/td[4]', 'tr[2]/td[4]', 'tr[1]/td[5]', 'tr[2]/td[5]']
AGM_LINK = 'https://www.agencymanager.amadeus.com/agm-connect-1.2.40/?configfile=/ahpqbtaub2f1f7fc'
ERROR = ['Missing value from table', 'Sales total not equal to 0', 'Comm or card not equal to 0']
errormsg = 0

def loadExcel():
    global EXCEL_PATH
    EXCEL_PATH = input('Please drag the excel workbook into this terminal:\n') #[1:-1]
    if EXCEL_PATH.startswith('"'):
        EXCEL_PATH = EXCEL_PATH[1:-1]
    
    while True:
        global excel
        excel = openpyxl.load_workbook(EXCEL_PATH)
        try:
            excel.save(EXCEL_PATH)
        except:
            print('Please close the excel sheet and/or ensure the path is correct.\n')
            EXCEL_PATH = input('Please drag the excel workbook into this terminal:\n')
            if EXCEL_PATH.startswith('"'):
                EXCEL_PATH = EXCEL_PATH[1:-1]
        else:
            break

    global sheet 
    sheet = excel.worksheets[0]

def windowCheck():
    for x in range(20):
        try:
            driver.find_element(By.XPATH, TABLE_XPATH + XPATH_VAR[0]).text
            TABLE_XPATH = TABLE_XPATH[:47] + str(x) + TABLE_XPATH[48:]
        except:
            continue

# Login check 
def loginCheck():
    a = 0
    while a < 120:
        if 'AGM Connect - QBT' in driver.title:
            try:
                driver.find_element(By.XPATH, '//*[@id="lblTokenkey"]')
            except:
                sleep(2)
                return
        else: 
            a = a + 2
            sleep(2)
    print('User Not logged in. Program will exit')
    exit()

# Checks sales coms and card are correct values
def diffCheck():
    try: 
        salesVal = float(driver.find_element(By.XPATH, TABLE_XPATH + XPATH_VAR[0]).text) + float(driver.find_element(By.XPATH, TABLE_XPATH + XPATH_VAR[1]).text)
    except:
        print(ERROR[0])
        errormsg = 0
        return False
    if int(salesVal) != 0: 
        print(ERROR[1])
        errormsg = 1
        return False
    for x in range(2, 6):
        try:
            value = float(driver.find_element(By.XPATH, TABLE_XPATH + XPATH_VAR[x]).text)
        except : 
            return False
            print(ERROR[0])
            errormsg = 0
        if int(value) != 0:
            return False
            print(ERROR[2])
            errormsg = 2
    return True

def acceptDiff(currRow):
    # right click to bring up menu 
    element = driver.find_element(By.XPATH, TABLE_XPATH + XPATH_VAR[0])
    action = ActionChains(driver)
    action.context_click(element).perform()
    sleep(0.3)

    # click on accept 
    accept = driver.find_element(By.XPATH, '//*[@id="PI_usr_Accept"]')
    accept.click()
    sleep(1)

    # click on ok 
    ok = driver.find_element(By.XPATH, '//*[@id="OK"]')
    ok.click()
    sleep(0.8)

    # get document number and input into excel 
    docnum = driver.find_element(By.ID, 'DocNum').get_attribute('value')
    sheet.cell(row = currRow, column = 4).value = docnum
    excel.save(EXCEL_PATH)

    # click cancel
    driver.find_element(By.XPATH, '//*[@id="Cancel"]').click()
    sleep(1)

###############################################################################
                        # START OF MAIN FUNCTION #
###############################################################################

loadExcel()

currRow = int(input("Enter current row of excel sheet, press ENTER if starting a new file : "))
if currRow == '':
    currRow = 3

# Load website
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
driver.get(AGM_LINK)

loginCheck()

# Get to the maintenance page 
action = ActionChains(driver)
action.key_down(Keys.CONTROL).send_keys(Keys.SPACE).key_up(Keys.CONTROL).perform()
action.send_keys('Maintenance of open differences').send_keys(Keys.RETURN).perform()
sleep(3)
action.send_keys(Keys.DOWN).perform()
sleep(1)
action.send_keys(Keys.RETURN).perform()
sleep(3)

again = 'y'
while again == 'y' or 'Y':
     
    while True:
        # Get ticket numbers
        tktnum = sheet.cell(row = currRow, column = 1).value

        if tktnum == sheet.cell(row = 1, column = 26).value:
            break

        # Input into agm and enter
        element = driver.find_element(By.ID, 'TicketNumber')
        element.send_keys(tktnum)
        action = ActionChains(driver)
        action.send_keys(Keys.RETURN).perform()
        sleep(1)

        # Check for errors 
        if diffCheck() == False:
            with open(EXCEL_PATH, "a") as f:
                f.write(str(tktnum) + ' ' + ERROR[errormsg] + '\n')
                print('Ticket Number: ' + str(tktnum) + ' has an error. Writing to file.')
        else:
            acceptDiff(currRow)
            acceptDiff(currRow + 1)

        # go to next ticket number 
        currRow = currRow + 2

        # remove ticket number after processing accepts
        driver.find_element(By.ID, 'TicketNumber').clear()

    print('File processing finished')
    again = input('Would you like to process another file? Press N for no Y for yes : \n')
    if again == 'y' or 'Y':
        loadExcel()
    else:
        exit()