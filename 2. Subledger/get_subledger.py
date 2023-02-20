from selenium.webdriver.chrome.webdriver import WebDriver as Chrome
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from time import sleep

from openpyxl import load_workbook
from openpyxl import utils
import pandas as pd

import os
import shutil

# Logging into AGM
APXNZ = "https://www.agencymanager.amadeus.com/agm-connect-1.2.40/?configfile=/ahpapxnz1e0af369"
USER_XPATH = "//*[@id='UserPrf']"
PASS_XPATH = "//*[@id='Password']"
LOGON_XPATH = "//*[@id='Logon']"
LOGON_TOKEN_XPATH = "//*[@id='edtToken']"
REQUEST_NEW_TOKEN_XPATH = "//*[@id='btnRequest']"
OK_XPATH = "//*[@id='OK']"
USERNAME = "user"
PASSWORD = "password"

# Inside APX NZ AGM (successful login)
MENU_XPATH = "//*[@id='tnFavMenu']/fieldset/agm-tab-bar/ul/div/div[2]/agm-tab-button"
BACK_OFFICE_XPATH = "//*[@id='MI222']/button"
OPEN_ITEMS_XPATH = "//*[@id='MI317']/button"
OPEN_ITEMS_CONSULTATION_XPATH = "//*[@id='MI318']/button"
GL_ACCOUNT_XPATH = "//*[@id='AccCde']"
SL_ACCOUNT_XPATH = "//*[@id='SLAcc']"
EXPORT_EXISTING_TO_CSV = "//*[@id='brw_1']/div[2]/agm-grid-button-icon[3]/button"
EXPORT_ALL_TO_CSV = "//*[@id='brw_1']/div[2]/agm-grid-button-icon[4]/button"
CONFIRM_EXPORT_ALL = "/html/body/ngb-modal-window/div/div/agm-message-box-content/div[3]/button[1]"
ALL_DATA_BUTTON = "//*[@id='brw_1']/div[1]/table/thead/tr/th[1]"
EXPORT_CANCEL_BUTTON = "/html/body/agm-root/div/agm-export-progress/div[1]/div/div/div[3]/button"
EXPORT_POPUP = "/html/body/agm-root/div/agm-export-progress/div[1]"

# get ascii files
OPEN_ITEMS_REPORT = "//*[@id='MI326']/button/span/span[1]"
OPEN_ITEMS_BY_REFERENCE = "//*[@id='MI327']/button/span/span[1]"
OPEN_ITEMS_GENERAL = "//*[@id='sys_tf']/fieldset/agm-tab-bar/ul/div/div[1]/agm-tab-button"
ACCOUNTANCY_PERIOD = "//*[@id='BkPerTo']"
GL_ACCOUNT1 = "//*[@id='AccCdeFrom']"
GL_ACCOUNT2 = "//*[@id='AccCdeTo']"
SL_ACCOUNT1 = "//*[@id='SLAccFrom']"
SL_ACCOUNT2 = "//*[@id='SLAccTo']"
SL_TYPE1 = "//*[@id='SLRelFrom']"
SL_TYPE2 = "//*[@id='SLRelTo']"
VALIDATE = "//*[@id='Validate']"
OK = "//*[@id='OK']"
EMAIL = "//*[@id='pnl_General']/fieldset/agm-dynamic/agm-check-box[2]/div/label"
CREATE_ASCII ="//*[@id='pnl_General']/fieldset/agm-dynamic/agm-check-box[4]/div"
VALIDATE2 = "//*[@id='WIN000005--Validate']"
OK2 = "//*[@id='WIN000005--OK']"

# Get ascii files
def get_df_ledgers(path):
    df = pd.read_excel(path, sheet_name="5080020000 ITBR") # change this
    GL_account = df.iat[0, 0]
    df = df.drop(df.index[:2])
    df.columns = df.iloc[0]
    df = df.drop(df.index[0])
    df.columns = ["Supplier No", "Supplier Name", "Currency", "Oct and prior", "Nov-22", "Dec-22", "Jan-23", "Total"] # change this
    df = df.iloc[:-5].reset_index(drop=True).fillna(0)
    df = df[(df["Oct and prior"] != 0) & (df["Supplier Name"] != "IATA BSP")].reset_index(drop=True)
    folder_path = "C:/Users/Matthew Chen/Desktop/subledger/{}_AGM_Files/".format(GL_account)
    SL_account_list = df[df['Oct and prior'] != 0]['Supplier No'].tolist()
    if not os.path.exists(folder_path):
        return GL_account, SL_account_list, df
    file_names = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
    temp = [x for x in file_names if x not in SL_account_list]
    temp2 = [int(x.split(".")[0].split("_")[1]) for x in temp]
    remaining_sl = []
    for sl in SL_account_list:
        if sl not in temp2:
            remaining_sl.append(sl)
    return GL_account, remaining_sl, df
def find_type_click_element(xpath, not_button=None, input=None, needs_enter=None, pause=None, wait=True, style=None):
    if wait:
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.visibility_of_element_located((By.XPATH, xpath)))
    element = driver.find_element(By.XPATH, xpath)
    if pause:
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
        action = ActionChains(driver)
        action.move_to_element(element).click().perform()
        sleep(1)
    else:
        element.click()
        # driver.execute_script('arguments[0].click();', element)
    if not_button:
        element.clear()
        element.send_keys(input)
    if needs_enter:
        element.send_keys(Keys.ENTER)
    return
def rename_move_file(general, subledger, is_first=True, to_be_updated=False, multiple=False, is_temp=False):
    desktop_path = ""
    initial_path = ""
    if is_first:
        is_first = False
        sleep(2)
    if to_be_updated:
        desktop_path = "C:/Users/Matthew Chen/Desktop/subledger/{}_AGM_Files/".format(general)
    else:
        # desktop_path = "C:/Users/ctm_mchen/OneDrive - Helloworld Travel Ltd/Desktop/General_Ledger_{}/".format(general)
        desktop_path = "C:/Users/Matthew Chen/Desktop/subledger/General_Ledger_{}/".format(general)
    if is_temp:
        desktop_path = "C:/Users/Matthew Chen/Desktop/subledger/{}_CSV_Files/".format(general)
        initial_path = "C:/temp/"
    else:
        # initial_path = "C:/Users/ctm_mchen/Downloads"
        initial_path = "C:/Users/Matthew Chen/Downloads"
    if not os.path.exists(desktop_path):
        os.mkdir(desktop_path)
    if multiple:
        agm_files = [f for f in os.listdir(initial_path) if f.endswith(".agm")]
        agm_filename = os.path.splitext(agm_files[0])[0]
        pdf_file = agm_filename + ".pdf"
        if pdf_file in os.listdir(initial_path):
            os.remove(os.path.join(initial_path, pdf_file))
        shutil.move(os.path.join(initial_path, agm_files[0]),
                    os.path.join(desktop_path, "Subledger_{}.agm".format(subledger))) 
    else:
        filename = max([initial_path + "/" + f for f in os.listdir(initial_path)],key=os.path.getctime)
        ext = os.path.splitext(filename)[1]
        shutil.move(filename,os.path.join(desktop_path,"Subledger_{}{}".format(subledger, ext)))
def initiate_driver():
    global driver
    options = Options()
    prefs = {'profile.default_content_setting_values.automatic_downloads': 1}
    options.add_experimental_option("prefs", prefs)
    driver = Chrome("chromedriver.exe", options=options)
    driver.maximize_window()
    driver.get(APXNZ)
    APXNZ_login()
def APXNZ_login():
    find_type_click_element(USER_XPATH, not_button=True, input=USERNAME)
    find_type_click_element(PASS_XPATH, not_button=True, input=PASSWORD)
    find_type_click_element(LOGON_XPATH)
    return
def get_updated_ascii(general, to_be_updated_subledgers):
    initiate_driver()
    for subledger in to_be_updated_subledgers:
        find_type_click_element(MENU_XPATH)
        find_type_click_element(BACK_OFFICE_XPATH)
        find_type_click_element(OPEN_ITEMS_XPATH)
        find_type_click_element(OPEN_ITEMS_REPORT)
        find_type_click_element(OPEN_ITEMS_BY_REFERENCE)
        find_type_click_element(OPEN_ITEMS_GENERAL)
        find_type_click_element(ACCOUNTANCY_PERIOD, not_button=True, input="202306")
        find_type_click_element(GL_ACCOUNT1, not_button=True, input=general)
        find_type_click_element(GL_ACCOUNT2, not_button=True, input=general)
        find_type_click_element(SL_ACCOUNT1, not_button=True, input=subledger)
        find_type_click_element(SL_ACCOUNT2, not_button=True, input=subledger)
        find_type_click_element(SL_TYPE1, not_button=True, input="1")
        find_type_click_element(SL_TYPE2, not_button=True, input="1")
        find_type_click_element(VALIDATE)
        find_type_click_element(OK)
        find_type_click_element(EMAIL)
        find_type_click_element(CREATE_ASCII)
        find_type_click_element(VALIDATE2, pause=True)
        find_type_click_element(OK2, pause=True)
        wait = WebDriverWait(driver, 120)
        wait.until(EC.invisibility_of_element_located((By.XPATH, OK2)))
        rename_move_file(general, subledger, to_be_updated=True, multiple=True)
    return

# read ascii files and get combined df
def append_all(dfs, names, df, name):
    dfs.append(df)
    names.append(name)
    return
def get_df_ledgers(path):
    df = pd.read_excel(path, sheet_name="5080020000 ITBR") # change this
    GL_account = df.iat[0, 0]
    df = df.drop(df.index[:2])
    df.columns = df.iloc[0]
    df = df.drop(df.index[0])
    df.columns = ["Supplier No", "Supplier Name", "Currency", "Oct and prior", "Nov-22", "Dec-22", "Jan-23", "Total"] # change this
    df = df.iloc[:-5].reset_index(drop=True).fillna(0)
    df = df[(df["Oct and prior"] != 0) & (df["Supplier Name"] != "IATA BSP")].reset_index(drop=True)
    folder_path = "C:/Users/Matthew Chen/Desktop/subledger/{}_AGM_Files/".format(GL_account)
    SL_account_list = df[df['Oct and prior'] != 0]['Supplier No'].tolist()
    if not os.path.exists(folder_path):
        return GL_account, SL_account_list, df
    file_names = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
    temp = [x for x in file_names if x not in SL_account_list]
    temp2 = [int(x.split(".")[0].split("_")[1]) for x in temp]
    remaining_sl = []
    for sl in SL_account_list:
        if sl not in temp2:
            remaining_sl.append(sl)
    return GL_account, remaining_sl, df
def read_agm_csv(filepath):
    ls = [0, 0, 0, 0, 0, 0, 0, 0, 0]
    with open(filepath, 'r') as file:
        for line in file:
            if line.startswith('D3'):
                temp = line.strip().split("\x1f")
                fy = temp[3]
                amount = float(temp[12].strip().replace(",", ""))
                if fy.startswith("2017"):
                    ls[0] += amount
                    ls[8] += amount
                elif fy.startswith("2018"):
                    ls[1] += amount
                    ls[8] += amount
                elif fy.startswith("2019"):
                    ls[2] += amount
                    ls[8] += amount
                elif fy.startswith("2020"):
                    ls[3] += amount
                    ls[8] += amount
                elif fy.startswith("2021"):
                    ls[4] += amount
                    ls[8] += amount
                elif fy.startswith("20220"):
                    ls[5] += amount
                    ls[8] += amount
                elif fy.startswith("20221"):
                    ls[6] += amount
                    ls[8] += amount
                elif fy == "202301" or fy == "202302" or fy == "202303" or fy == "202304":
                    ls[7] += amount
                    ls[8] += amount
    return ls
def get_combined_df(main_df, general):
    folder_path = "C:/Users/Matthew Chen/Desktop/subledger/{}_AGM_Files/".format(general)
    files = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
    files = sorted(files, key=lambda x: len(os.path.basename(x)))
    name_ls = []
    ls = []
    for file in files:
        name = file.split(".")[0].split("_")[1]
        path = folder_path + file
        name_ls.append(name)
        ls.append(read_agm_csv(path))
    fy_df = pd.DataFrame(ls, columns=['FY17', 'FY18', 'FY19', 'FY20', 'FY21', 'FY22 (to Mar22)', 'FY22 (Apr-Jun22)', 'FY23 (Jul-Oct22)', 'Total (Oct and prior)'])
    combined = pd.concat([main_df[['Supplier No', 'Supplier Name', 'Oct and prior']], fy_df], axis=1).fillna(0)
    combined = combined.assign(Diff = combined['Oct and prior'] - combined['Total (Oct and prior)'])
    append_all(dfs, names, combined, "Combined Summary")
    return

# add to workbook
def add_sheets(path):
    wb = load_workbook(path)
    ws = None
    for i in range(len(names)):
        ws = wb.create_sheet(names[i])
        for col_idx, value in enumerate(dfs[i].columns.tolist()):
            ws.cell(row=1, column=col_idx+1).value = value
        for row_idx, row in dfs[i].iterrows():
            for col_idx, cell_value in enumerate(row.tolist()):
                try:
                    rounded_value = round(float(cell_value), 2)
                    ws.cell(row=row_idx+2, column=col_idx+1).value = rounded_value
                except:
                    ws.cell(row=row_idx+2, column=col_idx+1).value = cell_value
        ws.freeze_panes = 'A2'
        for column in ws.columns:
            for cell in column:
                column_letter = utils.get_column_letter(cell.column)
                ws.column_dimensions[column_letter].auto_size = True
    wb.save(path)
    dfs.clear()
    names.clear()
    return


global dfs
global names
dfs = []
names = []
path = "APX ITBR Jan23.xlsx"

general, subledgers, main_df = get_df_ledgers(path)
combined = get_combined_df(main_df, general)
add_sheets(path)